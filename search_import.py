from datapackage_pipelines.wrapper import ingest, spew
from datapackage_pipelines.utilities.resources import PROP_STREAMING
from tabulator import Stream
import logging
from openpyxl import load_workbook


def get_sheet_files():
    with Stream('data/search_import/index.csv', headers=1) as stream:
        for sheet_file in stream.iter(keyed=True):
            yield sheet_file


def process_row(row):
    del_keys = []
    for k in row:
        if not k or not k.strip():
            del_keys.append(k)
        elif not row[k]:
            row[k] = ''
        else:
            row[k] = str(row[k])
    for k in del_keys:
        del row[k]
    return row


def sheet_iterator(first_row, header_row, stream_iter, filename, sheet_name, stream):
    def inner_sheet_iterator():
        if first_row:
            yield process_row(dict(zip(header_row, first_row)))
        for row in stream_iter:
            yield process_row(dict(zip(header_row, row)))
    for i, row in enumerate(inner_sheet_iterator()):
        if not row['migdar_id']:
            if row['title']:
                logging.warning('missing migdar_id for {}/{} row number {}'.format(filename, sheet_name, i))
        else:
            yield row
    stream.close()


def load_sheets():
    schema = {'fields': []}
    sheets = []
    for sheet_file in get_sheet_files():
        filename = 'data/search_import/{}'.format(sheet_file['name'])
        wb = load_workbook(filename)
        for sheet_number, sheet_name in enumerate(wb.sheetnames, start=1):
            stream = Stream(filename, sheet=sheet_name)
            stream.open()
            # logging.info('{}/{}'.format(filename, sheet_name))
            stream_iter = stream.iter()
            first_row = next(stream_iter)
            if 'migdar_id' not in first_row and sheet_number > 1:
                header_row = first_sheet_header_row
            else:
                header_row = first_row
                first_row = None
                if sheet_number == 1:
                    first_sheet_header_row = header_row
            # logging.info(header_row)
            # logging.info(first_row)
            for k in header_row:
                if k and k not in [f['name'] for f in schema['fields']]:
                    # logging.info('found field: {}'.format(k))
                    field_type = 'string'
                    schema['fields'].append({'name': k, 'type': field_type})
            sheets.append({'iterator': sheet_iterator(first_row, header_row, stream_iter, filename, sheet_name, stream),
                           'deleted': 'deleted' in sheet_name.strip().lower()})
    return schema, sheets


def get_datapackage_resources(schema):
    return [{PROP_STREAMING: True, "name": "cataloged", "path": "cataloged.csv",
             "schema": schema},
            {PROP_STREAMING: True, 'name': 'deleted', 'path': 'deleted.csv',
             'schema': {'fields': [{'name': 'migdar_id', 'type': 'string'}]}}]


def get_resource(deleted, sheets):
    for sheet in sheets:
        # logging.info(sheet)
        # logging.info('deleted={}'.format(deleted))
        if sheet['deleted'] == deleted:
            for row in sheet['iterator']:
                if deleted:
                    # logging.info(row)
                    yield {'migdar_id': str(row['migdar_id'])}
                else:
                    yield row


def main():
    parameters, datapackage, resources, stats = ingest() + ({},)
    schema, sheets = load_sheets()
    datapackage["resources"] = get_datapackage_resources(schema)
    spew(datapackage, [get_resource(deleted=False, sheets=sheets), get_resource(deleted=True, sheets=sheets)], stats)


if __name__ == '__main__':
    main()
